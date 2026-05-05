import csv
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from database.db import get_db
from database.models import PipeSchedule
from schemas.pipe_table import PipeScheduleOut

router = APIRouter(prefix="/api/pipe-schedule", tags=["pipe-schedule"])

# DN → NPS 매핑 (ASME B36.10/B36.19)
_DN_TO_NPS = {
    6: 0.125, 8: 0.25, 10: 0.375,
    15: 0.5, 20: 0.75, 25: 1.0, 32: 1.25, 40: 1.5,
    50: 2.0, 65: 2.5, 80: 3.0, 90: 3.5,
    100: 4.0, 125: 5.0, 150: 6.0,
    200: 8.0, 250: 10.0, 300: 12.0, 350: 14.0,
    400: 16.0, 450: 18.0, 500: 20.0, 550: 22.0,
    600: 24.0, 650: 26.0, 700: 28.0, 750: 30.0,
    800: 32.0, 850: 34.0, 900: 36.0,
}


@router.get("", response_model=list[PipeScheduleOut])
def get_pipe_schedules(
    standard: str | None = Query(None),
    dn: int | None = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(PipeSchedule)
    if standard:
        q = q.filter(PipeSchedule.standard == standard)
    if dn is not None:
        q = q.filter(PipeSchedule.dn == dn)
    return q.order_by(PipeSchedule.id).all()


@router.get("/export")
def export_pipe_schedules(
    standard: str | None = Query(None),
    db: Session = Depends(get_db),
):
    import openpyxl

    q = db.query(PipeSchedule)
    if standard:
        q = q.filter(PipeSchedule.standard == standard)
    rows = q.order_by(PipeSchedule.standard, PipeSchedule.dn, PipeSchedule.wt_mm).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "pipe_schedule"
    ws.append([
        "standard", "dn", "nps", "schedule", "identification",
        "od_mm", "wt_mm", "mass_kg_m", "od_in", "wt_in", "mass_lb_ft",
    ])
    for r in rows:
        ws.append([
            r.standard, r.dn, r.nps, r.schedule, r.identification,
            r.od_mm, r.wt_mm, r.mass_kg_m, r.od_in, r.wt_in, r.mass_lb_ft,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"pipe_schedule_{standard or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _insert_rows(db: Session, rows: list[dict]) -> tuple[int, int]:
    inserted = skipped = 0
    for r in rows:
        obj = PipeSchedule(**r)
        db.add(obj)
        try:
            db.commit()
            inserted += 1
        except IntegrityError:
            db.rollback()
            skipped += 1
    return inserted, skipped


def _parse_csv(content: str) -> list[dict]:
    reader = csv.DictReader(io.StringIO(content))
    rows = []
    for row in reader:
        rows.append({
            "standard": row["standard"].strip(),
            "dn": int(row["dn"]),
            "nps": float(row["nps"]),
            "schedule": row["schedule"].strip(),
            "identification": row.get("identification") or None,
            "od_mm": float(row["od_mm"]),
            "wt_mm": float(row["wt_mm"]),
            "mass_kg_m": float(row["mass_kg_m"]) if row.get("mass_kg_m") else None,
            "od_in": float(row["od_in"]) if row.get("od_in") else None,
            "wt_in": float(row["wt_in"]) if row.get("wt_in") else None,
            "mass_lb_ft": float(row["mass_lb_ft"]) if row.get("mass_lb_ft") else None,
        })
    return rows


def _parse_excel_b3610(ws) -> list[dict]:
    """B36.10 시트 파싱.
    헤더: NO, NPS, Customary_OD_in, Customary_WT_in, Customary_Mass_lbft,
          Identification, SCH, DN, SI_OD_mm, SI_WT_mm, SI_Mass_kgm
    """
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        _, _, od_in, wt_in, mass_lbft, ident, sch, dn, od_mm, wt_mm, mass_kgm = row
        if dn is None or wt_mm is None:
            continue
        if sch is not None:
            schedule = f"SCH {sch}"
        elif ident is not None:
            schedule = ident
        else:
            schedule = None
        rows.append({
            "standard": "B36.10",
            "dn": int(dn),
            "nps": _DN_TO_NPS.get(int(dn), round(float(od_in or 0), 3)),
            "schedule": schedule,
            "identification": ident or None,
            "od_mm": float(od_mm),
            "wt_mm": float(wt_mm),
            "mass_kg_m": float(mass_kgm) if mass_kgm is not None else None,
            "od_in": float(od_in) if od_in is not None else None,
            "wt_in": float(wt_in) if wt_in is not None else None,
            "mass_lb_ft": float(mass_lbft) if mass_lbft is not None else None,
        })
    return rows


def _parse_excel_b3619(ws) -> list[dict]:
    """B36.19 시트 파싱.
    헤더: NO, NPS, Customary_OD_in, Customary_WT_in, Customary_Mass_lbft,
          SCH, DN, SI_OD_mm, SI_WT_mm, SI_Mass_kgm, Remark
    """
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        _, _, od_in, wt_in, mass_lbft, sch, dn, od_mm, wt_mm, mass_kgm, _ = row
        if dn is None or wt_mm is None or sch is None:
            continue
        rows.append({
            "standard": "B36.19",
            "dn": int(dn),
            "nps": _DN_TO_NPS.get(int(dn), round(float(od_in or 0), 3)),
            "schedule": f"SCH {sch}",
            "identification": None,
            "od_mm": float(od_mm),
            "wt_mm": float(wt_mm),
            "mass_kg_m": float(mass_kgm) if mass_kgm is not None else None,
            "od_in": float(od_in) if od_in is not None else None,
            "wt_in": float(wt_in) if wt_in is not None else None,
            "mass_lb_ft": float(mass_lbft) if mass_lbft is not None else None,
        })
    return rows


def _parse_excel_simple(ws) -> list[dict]:
    """다운로드 후 재업로드 용 단순 포맷 파싱.
    헤더: standard, dn, nps, schedule, identification,
          od_mm, wt_mm, mass_kg_m, od_in, wt_in, mass_lb_ft
    """
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    col = {h: i for i, h in enumerate(header) if h}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dn = row[col["dn"]] if "dn" in col else None
        wt_mm = row[col["wt_mm"]] if "wt_mm" in col else None
        if dn is None or wt_mm is None:
            continue

        def g(key):
            return row[col[key]] if key in col else None

        rows.append({
            "standard": str(g("standard") or "").strip(),
            "dn": int(dn),
            "nps": float(g("nps") or 0),
            "schedule": str(g("schedule")).strip() if g("schedule") is not None else None,
            "identification": str(g("identification")).strip() if g("identification") is not None else None,
            "od_mm": float(g("od_mm")),
            "wt_mm": float(wt_mm),
            "mass_kg_m": float(g("mass_kg_m")) if g("mass_kg_m") is not None else None,
            "od_in": float(g("od_in")) if g("od_in") is not None else None,
            "wt_in": float(g("wt_in")) if g("wt_in") is not None else None,
            "mass_lb_ft": float(g("mass_lb_ft")) if g("mass_lb_ft") is not None else None,
        })
    return rows


def _parse_excel(content: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 패키지가 필요합니다: uv add openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(content))
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            header = [c.value for c in next(ws.iter_rows(max_row=1))]
        except StopIteration:
            continue
        if "standard" in header:
            rows.extend(_parse_excel_simple(ws))
        elif "Identification" in header:
            rows.extend(_parse_excel_b3610(ws))
        elif "Remark" in header:
            rows.extend(_parse_excel_b3619(ws))
    if not rows:
        raise HTTPException(status_code=400, detail="인식할 수 없는 Excel 형식입니다. B36.10/B36.19 양식 또는 내보낸 파일을 사용하세요.")
    return rows


@router.post("/upload")
def upload_pipe_file(
    file: UploadFile = File(...),
    mode: str = Query("add", pattern="^(add|replace)$"),
    db: Session = Depends(get_db),
):
    raw = file.file.read()
    fname = (file.filename or "").lower()

    if fname.endswith(".xlsx") or fname.endswith(".xls"):
        rows = _parse_excel(raw)
    else:
        rows = _parse_csv(raw.decode("utf-8-sig"))

    if mode == "replace":
        db.query(PipeSchedule).delete()
        db.commit()

    inserted, skipped = _insert_rows(db, rows)
    return {"inserted": inserted, "skipped": skipped}


@router.delete("/{item_id}")
def delete_pipe_schedule(item_id: int, db: Session = Depends(get_db)):
    obj = db.query(PipeSchedule).filter(PipeSchedule.id == item_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(obj)
    db.commit()
    return {"deleted": item_id}
