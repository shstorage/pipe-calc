import csv
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from database.db import get_db
from database.models import AllowableStress
from schemas.allowable_stress import AllowableStressOut

router = APIRouter(prefix="/api/allowable-stress", tags=["allowable-stress"])

_COLS = ["code", "edition", "spec_no", "grade", "type_or_class", "nominal_comp", "p_no",
         "temp_c", "stress_mpa", "is_creep"]


@router.get("/meta")
def get_stress_meta(db: Session = Depends(get_db)):
    """DB에 실제 존재하는 코드·년판·Spec 목록 반환 (필터 드롭다운용)."""
    editions = sorted(
        {r[0] for r in db.query(AllowableStress.edition).distinct().all() if r[0]},
        reverse=True,
    )
    codes = sorted(
        {r[0] for r in db.query(AllowableStress.code).distinct().all() if r[0]}
    )
    specs = sorted(
        {r[0] for r in db.query(AllowableStress.spec_no).distinct().all() if r[0]}
    )
    return {"codes": codes, "editions": editions, "specs": specs}


@router.get("", response_model=list[AllowableStressOut])
def get_allowable_stresses(
    code: str | None = Query(None),
    edition: str | None = Query(None),
    spec_no: str | None = Query(None),
    grade: str | None = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(AllowableStress)
    if code:
        q = q.filter(AllowableStress.code == code)
    if edition:
        q = q.filter(AllowableStress.edition == edition)
    if spec_no:
        q = q.filter(AllowableStress.spec_no == spec_no)
    if grade:
        q = q.filter(AllowableStress.grade.ilike(f"%{grade}%"))
    return q.order_by(
        AllowableStress.code, AllowableStress.spec_no,
        AllowableStress.grade, AllowableStress.type_or_class,
        AllowableStress.temp_c,
    ).all()


@router.get("/template")
def download_template():
    """빈 Excel 양식 다운로드 — 편집 후 업로드하면 대량 추가 가능."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 패키지가 필요합니다.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "allowable_stress"
    ws.append(_COLS)
    ws.append(["B31.1", "2022", "A106", "B", "", "C-Si", "1", 37.78, 117.9, 0])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=allowable_stress_template.xlsx"},
    )


@router.get("/export")
def export_allowable_stresses(
    code: str | None = Query(None),
    edition: str | None = Query(None),
    db: Session = Depends(get_db),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 패키지가 필요합니다.")

    q = db.query(AllowableStress)
    if code:
        q = q.filter(AllowableStress.code == code)
    if edition:
        q = q.filter(AllowableStress.edition == edition)
    rows = q.order_by(
        AllowableStress.code, AllowableStress.edition,
        AllowableStress.spec_no, AllowableStress.grade,
        AllowableStress.type_or_class, AllowableStress.temp_c,
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "allowable_stress"
    ws.append(_COLS)
    for r in rows:
        ws.append([r.code, r.edition or "", r.spec_no, r.grade,
                   r.type_or_class or "", r.nominal_comp or "", r.p_no or "",
                   r.temp_c, r.stress_mpa, r.is_creep])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    parts = [p for p in [code, edition] if p]
    fname = f"allowable_stress_{'_'.join(parts) if parts else 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _parse_row(r: dict) -> dict | None:
    code = str(r.get("code") or "").strip()
    spec_no = str(r.get("spec_no") or "").strip()
    grade = str(r.get("grade") or "").strip()
    temp_c = r.get("temp_c")
    stress_mpa = r.get("stress_mpa")
    if not (code and spec_no and grade and temp_c is not None and stress_mpa is not None):
        return None
    try:
        return {
            "code": code,
            "edition": str(r.get("edition") or "").strip() or None,
            "spec_no": spec_no,
            "grade": grade,
            "type_or_class": str(r.get("type_or_class") or "").strip(),
            "nominal_comp": str(r.get("nominal_comp") or "").strip() or None,
            "p_no": str(r.get("p_no") or "").strip() or None,
            "temp_c": float(temp_c),
            "stress_mpa": float(stress_mpa),
            "is_creep": int(r.get("is_creep") or 0),
        }
    except (ValueError, TypeError):
        return None


def _parse_csv(content: str) -> list[dict]:
    reader = csv.DictReader(io.StringIO(content))
    return [r for raw in reader if (r := _parse_row(raw))]


def _parse_excel(content: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 패키지가 필요합니다.")

    wb = openpyxl.load_workbook(io.BytesIO(content))
    results = []
    for ws in wb.worksheets:
        try:
            header = [c.value for c in next(ws.iter_rows(max_row=1))]
        except StopIteration:
            continue
        if "spec_no" not in header or "stress_mpa" not in header:
            continue
        col = {h: i for i, h in enumerate(header) if h}
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = {k: row[v] for k, v in col.items()}
            parsed = _parse_row(raw)
            if parsed:
                results.append(parsed)
    if not results:
        raise HTTPException(
            status_code=400,
            detail="인식할 수 없는 형식입니다. '양식 다운로드' 파일을 기준으로 작성하세요.",
        )
    return results


def _insert_rows(db: Session, rows: list[dict]) -> tuple[int, int]:
    inserted = skipped = 0
    for r in rows:
        obj = AllowableStress(**r)
        db.add(obj)
        try:
            db.commit()
            inserted += 1
        except IntegrityError:
            db.rollback()
            skipped += 1
    return inserted, skipped


@router.post("/upload")
def upload_stress_file(
    file: UploadFile = File(...),
    mode: str = Query("add", pattern="^(add|replace)$"),
    db: Session = Depends(get_db),
):
    if mode == "replace":
        db.query(AllowableStress).delete()
        db.commit()

    raw = file.file.read()
    fname = (file.filename or "").lower()

    if fname.endswith(".xlsx") or fname.endswith(".xls"):
        rows = _parse_excel(raw)
    else:
        rows = _parse_csv(raw.decode("utf-8-sig"))

    inserted, skipped = _insert_rows(db, rows)
    return {"inserted": inserted, "skipped": skipped}


@router.delete("/{item_id}")
def delete_allowable_stress(item_id: int, db: Session = Depends(get_db)):
    obj = db.query(AllowableStress).filter(AllowableStress.id == item_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(obj)
    db.commit()
    return {"deleted": item_id}
